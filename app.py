import streamlit as st
from datetime import datetime, date
import pandas as pd
import calendar, requests, json
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl.styles import Font, Alignment

st.set_page_config(page_title="Livro de Ponto", page_icon="⛪", layout="wide")

SCRIPT_URL = "https://script.google.com/macros/s/AKfycbzK268XUzyLr4TJlksHJf1k8GEPyOG7DrwItH8iw_dSdb6ysVShyEQlsau2lth27kKY/exec"
SHEET_ID = "1KhoFXD3oB91U-gh1zy1-YCK4Kug4XfdlpwttXPU6KAY"
COLS = ["Data","Dia da Semana","Entrada","Saída","Horas Trabalhadas","Atraso","Desconto (MZN)","Notas"]
VALOR_HORA = 45.82
HORA_LIMITE = "08:40"
HORA_MAX_SAIDA = "16:40"

feriados = {
    "01-01":"Ano Novo","03-02":"Dia dos Heróis Moçambicanos",
    "07-04":"Dia da Mulher Moçambicana","01-05":"Dia do Trabalhador",
    "25-06":"Dia da Independência Nacional","07-09":"Dia da Vitória",
    "25-09":"Dia das Forças Armadas","04-10":"Paz e Reconciliação",
    "25-12":"Dia da Família"
}
meses = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",
         6:"Junho",7:"Julho",8:"Agosto",9:"Setembro",
         10:"Outubro",11:"Novembro",12:"Dezembro"}
dias_pt = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']

def clean(x):
    s = str(x)
    return "" if s in ["nan","None","NaN",""] else s

def calc_horas(ent, sai):
    try:
        e = datetime.strptime(ent.strip(), "%H:%M")
        s = datetime.strptime(sai.strip(), "%H:%M")
        smax = datetime.strptime(HORA_MAX_SAIDA, "%H:%M")
        if s > smax:
            s = smax
        d = int((s - e).total_seconds() // 60) - 60
        if d <= 0:
            return "Erro: tempo insuficiente", False
        return f"{d//60}h {d%60:02d}m", True
    except:
        return "Formato inválido (use HH:MM)", False

def calc_atraso(ent):
    try:
        e = datetime.strptime(ent.strip(), "%H:%M")
        lim = datetime.strptime(HORA_LIMITE, "%H:%M")
        if e <= lim:
            return "Sem atraso", 0.0
        min_atr = int((e - lim).total_seconds() // 60)
        desc = round((min_atr / 60) * VALOR_HORA, 2)
        h = min_atr // 60
        m = min_atr % 60
        atr_str = f"{h}h {m:02d}m" if h > 0 else f"{m} min"
        return atr_str, desc
    except:
        return "Erro", 0.0

def tot_min(df):
    t = 0
    for h in df["Horas Trabalhadas"]:
        h = str(h)
        if "h" in h:
            p = h.replace("m","").split("h")
            try:
                t += int(p[0])*60 + int(p[1].strip())
            except:
                pass
    return t

def tot_desc(df):
    t = 0.0
    if "Desconto (MZN)" in df.columns:
        for v in df["Desconto (MZN)"]:
            try:
                v = str(v).replace(",",".")
                if v not in ["nan","None","","Sem desconto","0","0.0"]:
                    t += float(v)
            except:
                pass
    return round(t, 2)

@st.cache_data(ttl=30)
def load():
    try:
        return pd.read_csv(
            f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv"
        )
    except:
        return pd.DataFrame(columns=COLS+["Mês","Ano"])

def save(reg):
    try:
        requests.post(SCRIPT_URL, data=json.dumps(reg),
            headers={"Content-Type":"application/json"},
            timeout=15, allow_redirects=True)
        return True
    except Exception as e:
        st.error(str(e))
        return False
def make_pdf(df, mes_nome, ano, tot_h, tot_m, tot_d):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    S = getSampleStyleSheet()
    el = []
    t1 = ParagraphStyle('t1',parent=S['Normal'],fontSize=16,fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=6)
    t2 = ParagraphStyle('t2',parent=S['Normal'],fontSize=13,fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=4)
    ti = ParagraphStyle('ti',parent=S['Normal'],fontSize=11,fontName='Helvetica',alignment=TA_CENTER,spaceAfter=2)
    tn = ParagraphStyle('tn',parent=S['Normal'],fontSize=10,fontName='Helvetica',alignment=TA_LEFT,spaceAfter=2)
    tf = ParagraphStyle('tf',parent=S['Normal'],fontSize=10,fontName='Helvetica',alignment=TA_CENTER,spaceAfter=2)
    tr = ParagraphStyle('tr',parent=S['Normal'],fontSize=10,fontName='Helvetica-Bold',alignment=TA_CENTER,textColor=colors.HexColor("#c0392b"),spaceAfter=2)
    el.append(Paragraph("Paróquia SS. Trindade", t1))
    el.append(Paragraph("Livro de Ponto — Yolanda Facitela Clávio", t2))
    el.append(Spacer(1, 0.2*cm))
    el.append(Paragraph(f"Mês: {mes_nome} {ano}", ti))
    el.append(Spacer(1, 0.2*cm))
    resumo = [[
        Paragraph(f"<b>Total Horas:</b> {tot_h}h {tot_m:02d}m", ti),
        Paragraph(f"<b>Total Desconto Atrasos:</b> {tot_d:.2f} MZN", tr)
    ]]
    tr2 = Table(resumo, colWidths=[8.5*cm,8.5*cm])
    tr2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    el.append(tr2)
    el.append(Spacer(1, 0.3*cm))
    el.append(Table([['']],colWidths=[17*cm],style=TableStyle([('LINEBELOW',(0,0),(-1,-1),1,colors.black)])))
    el.append(Spacer(1, 0.4*cm))
    rows = [["Data","Dia","Entrada","Saída","Horas","Atraso","Desconto\n(MZN)","Notas"]]
    for _, r in df.iterrows():
        rows.append([
            clean(r.get("Data","")), clean(r.get("Dia da Semana","")),
            clean(r.get("Entrada","")), clean(r.get("Saída","")),
            clean(r.get("Horas Trabalhadas","")), clean(r.get("Atraso","")),
            clean(r.get("Desconto (MZN)","")), clean(r.get("Notas",""))
        ])
    tab = Table(rows, colWidths=[2.3*cm,2.4*cm,1.8*cm,1.8*cm,1.8*cm,1.8*cm,2.0*cm,3.1*cm], repeatRows=1)
    tab.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#2c3e50")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor("#f2f2f2")]),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('TOPPADDING',(0,0),(-1,-1),4),
        ('BOTTOMPADDING',(0,0),(-1,-1),4),
    ]))
    el.append(tab)
    el.append(Spacer(1, 1.5*cm))
    ass = [
        [Paragraph("_______________________________",tf), Paragraph("_______________________________",tf)],
        [Paragraph("<b>Pároco: Pe. Pasquale Peluso</b>",tf), Paragraph("<b>Secretária: Yolanda Facitela Clávio</b>",tf)],
        [Paragraph("Data: ___/___/______",tf), Paragraph("Data: ___/___/______",tf)]
    ]
    ta = Table(ass, colWidths=[8.5*cm,8.5*cm])
    ta.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('TOPPADDING',(0,0),(-1,-1),8)]))
    el.append(Paragraph("Assinaturas:", tn))
    el.append(Spacer(1, 0.3*cm))
    el.append(ta)
    doc.build(el)
    buf.seek(0)
    return buf

def make_excel(df, mes_nome, ano, tot_h, tot_m, tot_d):
    out = BytesIO()
    cols_exp = [c for c in COLS if c in df.columns] if not df.empty else COLS
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        d = df[cols_exp].copy() if not df.empty else pd.DataFrame(columns=cols_exp)
        for col in d.columns:
            d[col] = d[col].apply(clean)
        d.to_excel(w, sheet_name=mes_nome, index=False, startrow=8)
        ws = w.sheets[mes_nome]
        ws["A1"] = "Paróquia SS. Trindade"
        ws["A2"] = "Livro de Ponto — Yolanda Facitela Clávio"
        ws["A3"] = "Pároco: Pe. Pasquale Peluso"
        ws["A4"] = "Secretária: Yolanda Facitela Clávio"
        ws["A5"] = f"Mês: {mes_nome} {ano}"
        ws["A6"] = f"Total Horas Trabalhadas: {tot_h}h {tot_m:02d}m"
        ws["A7"] = f"Total Desconto Atrasos: {tot_d:.2f} MZN"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"].font = Font(bold=True, size=12)
        ws["A6"].font = Font(bold=True)
        ws["A7"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"].alignment = Alignment(horizontal="center")
    return out.getvalue()
# SIDEBAR
st.sidebar.title("📅 Meses")
ano = datetime.now().year
mes = st.sidebar.selectbox("Selecione o mês:",
    options=list(meses.keys()),
    format_func=lambda x: meses[x],
    index=datetime.now().month-1)
st.sidebar.markdown(f"### {meses[mes]} {ano}")
st.sidebar.markdown("---")
num_dias = calendar.monthrange(ano, mes)[1]
df0 = load()
if not df0.empty and "Mês" in df0.columns:
    dm0 = df0[(df0["Mês"].astype(str)==str(mes))&(df0["Ano"].astype(str)==str(ano))]
    if not dm0.empty:
        t0 = tot_min(dm0)
        td0 = tot_desc(dm0)
        st.sidebar.success(f"⏱️ **Total Horas:** {t0//60}h {t0%60:02d}m")
        st.sidebar.info(f"📅 **Dias:** {len(dm0)}")
        if td0 > 0:
            st.sidebar.error(f"💸 **Desconto:** {td0:.2f} MZN")
    else:
        st.sidebar.info("Nenhum registo este mês")
st.title("⛪ Paróquia SS. Trindade")
st.subheader("Livro de Ponto — Yolanda Facitela Clávio")
st.markdown("**Pároco:** Pe. Pasquale Peluso | **Secretária:** Yolanda Facitela Clávio")
st.markdown("---")

st.subheader(f"📝 Novo Registo — {meses[mes]} {ano}")
c1, c2, c3 = st.columns(3)

with c1:
    dia = st.selectbox("Dia do mês",
        options=list(range(1, num_dias+1)),
        index=min(datetime.now().day-1, num_dias-1)
              if mes==datetime.now().month else 0,
        key=f"d_{mes}")
    data_obj = date(ano, mes, dia)
    dsem = dias_pt[data_obj.weekday()]
    st.info(f"Data: {data_obj.strftime('%d/%m/%Y')} | {dsem}")
    fer = feriados.get(data_obj.strftime("%d-%m"), "")
    if fer:
        st.warning(f"🎉 {fer}")

with c2:
    entrada = st.text_input("⏰ Hora de Entrada",
        placeholder="08:00", key=f"e_{mes}_{dia}")
    if entrada:
        atr, desc = calc_atraso(entrada)
        if desc > 0:
            st.error(f"Atraso: {atr} | Desconto: {desc:.2f} MZN")
        else:
            st.success("Entrada no horário")

with c3:
    saida = st.text_input("⏰ Hora de Saída",
        placeholder="16:30", key=f"s_{mes}_{dia}")

notas = st.text_input("📝 Notas", value="",
    placeholder="Escreva aqui se necessário...",
    key=f"n_{mes}_{dia}")
st.caption("ℹ️ Pausa almoço (13h-14h) descontada. Atraso a partir das 08h40.")

if st.button("✅ Guardar Registo", type="primary", use_container_width=True):
    if entrada and saida:
        horas, ok = calc_horas(entrada, saida)
        if ok:
            atr, desc = calc_atraso(entrada)
            reg = {
                "Data": data_obj.strftime("%d/%m/%Y"),
                "DiaSemana": dsem,
                "Entrada": entrada,
                "Saida": saida,
                "Horas": horas,
                "Atraso": atr,
                "Desconto": str(desc),
                "Notas": notas,
                "Mes": mes,
                "Ano": ano
            }
            with st.spinner("A guardar..."):
                if save(reg):
                    st.success(f"✅ Guardado! {data_obj.strftime('%d/%m/%Y')} — {horas}")
                    st.balloons()
                    st.cache_data.clear()
        else:
            st.error(horas)
    else:
        st.warning("⚠️ Insira a hora de entrada e saída.")

st.markdown("---")
st.subheader(f"📊 Registos de {meses[mes]} {ano}")
df_all = load()

if not df_all.empty and "Mês" in df_all.columns:
    do_mes = df_all[
        (df_all["Mês"].astype(str)==str(mes)) &
        (df_all["Ano"].astype(str)==str(ano))
    ].copy()
else:
    do_mes = pd.DataFrame(columns=COLS)

if not do_mes.empty:
    do_mes = do_mes.sort_values("Data")
    for col in do_mes.columns:
        do_mes[col] = do_mes[col].apply(clean)
    cols_ok = [c for c in COLS if c in do_mes.columns]
    st.dataframe(do_mes[cols_ok], use_container_width=True, hide_index=True)
    tot = tot_min(do_mes)
    td = tot_desc(do_mes)
    m1, m2, m3 = st.columns(3)
    m1.metric("⏱️ Total Horas", f"{tot//60}h {tot%60:02d}m")
    m2.metric("📅 Dias Trabalhados", len(do_mes))
    m3.metric("💸 Total Desconto", f"{td:.2f} MZN")
else:
    do_mes = pd.DataFrame(columns=COLS)
    tot = 0
    td = 0.0
    st.info(f"Nenhum registo para {meses[mes]} {ano}.")

st.markdown("---")
st.subheader("📥 Exportar para Assinar")
b1, b2 = st.columns(2)

with b1:
    st.download_button(
        label=f"📄 Baixar PDF — {meses[mes]} {ano}",
        data=make_pdf(do_mes, meses[mes], ano, tot//60, tot%60, td),
        file_name=f"LivroPonto_{meses[mes]}_{ano}.pdf",
        mime="application/pdf",
        use_container_width=True
    )

with b2:
    st.download_button(
        label=f"📊 Baixar Excel — {meses[mes]} {ano}",
        data=make_excel(do_mes, meses[mes], ano, tot//60, tot%60, td),
        file_name=f"LivroPonto_{meses[mes]}_{ano}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.info("💡 Abra o PDF ou Excel, imprima e entregue para assinatura da Yolanda.")
st.markdown("---")
st.caption("Paróquia SS. Trindade — Maputo, Moçambique")

