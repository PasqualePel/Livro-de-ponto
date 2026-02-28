from datetime import datetime
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl.styles import Font, Alignment

feriados = {
    "01-01": "Ano Novo",
    "03-02": "Dia dos Heróis Moçambicanos",
    "07-04": "Dia da Mulher Moçambicana",
    "01-05": "Dia do Trabalhador",
    "25-06": "Dia da Independência Nacional",
    "07-09": "Dia da Vitória",
    "25-09": "Dia das Forças Armadas",
    "04-10": "Paz e Reconciliação",
    "25-12": "Dia da Família"
}

meses = {
    1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril",
    5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto",
    9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"
}

dias_semana_pt = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']

def calc_horas(ent, sai):
    try:
        e = datetime.strptime(ent.strip(), "%H:%M")
        s = datetime.strptime(sai.strip(), "%H:%M")
        diff = int((s - e).total_seconds() // 60)
        if diff <= 0:
            return "Erro: saída antes da entrada", False
        return f"{diff//60}h {diff%60:02d}m", True
    except:
        return "Formato inválido (use HH:MM)", False

def gerar_pdf(do_mes, mes_nome, ano, total_h, total_m):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    styles = getSampleStyleSheet()
    elementos = []

    t1 = ParagraphStyle('t1', parent=styles['Normal'],
        fontSize=16, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=6)
    t2 = ParagraphStyle('t2', parent=styles['Normal'],
        fontSize=13, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=4)
    ti = ParagraphStyle('ti', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica',
        alignment=TA_CENTER, spaceAfter=2)
    tn = ParagraphStyle('tn', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica',
        alignment=TA_LEFT, spaceAfter=2)
    tf = ParagraphStyle('tf', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica',
        alignment=TA_CENTER, spaceAfter=2)

    elementos.append(Paragraph("Paróquia SS. Trindade", t1))
    elementos.append(Paragraph("Livro de Ponto — Yolanda Facitela Clávio", t2))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(f"Mês: {mes_nome} {ano}", ti))
    elementos.append(Paragraph(
        f"Total de Horas Trabalhadas: {total_h}h {total_m:02d}m", ti))
    elementos.append(Spacer(1, 0.5*cm))

    elementos.append(Table(
        [['']],
        colWidths=[17*cm],
        style=TableStyle([('LINEBELOW',(0,0),(-1,-1),1,colors.black)])
    ))
    elementos.append(Spacer(1, 0.5*cm))

    cabecalho = ["Data","Dia da Semana","Entrada","Saída","Horas","Notas"]
    dati = [cabecalho]
    for _, row in do_mes.iterrows():
        dati.append([
            str(row.get("Data","")),
            str(row.get("Dia da Semana","")),
            str(row.get("Entrada","")),
            str(row.get("Saída","")),
            str(row.get("Horas Trabalhadas","")),
            str(row.get("Notas",""))
        ])

    tab = Table(dati,
        colWidths=[2.5*cm,3.2*cm,2.2*cm,2.2*cm,2.2*cm,4.7*cm],
        repeatRows=1)
    tab.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#2c3e50")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor("#f2f2f2")]),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('TOPPADDING',(0,0),(-1,-1),5),
    ]))
    elementos.append(tab)
    elementos.append(Spacer(1, 2*cm))

    ass = [
        [Paragraph("_______________________________",tf),
         Paragraph("_______________________________",tf)],
        [Paragraph("<b>Pároco: Pe. Pasquale Peluso</b>",tf),
         Paragraph("<b>Secretária: Yolanda Facitela Clávio</b>",tf)],
        [Paragraph("Data: _____ / _____ / _________",tf),
         Paragraph("Data: _____ / _____ / _________",tf)]
    ]
    tab_ass = Table(ass, colWidths=[8.5*cm,8.5*cm])
    tab_ass.setStyle(TableStyle([
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),8),
    ]))
    elementos.append(Paragraph("Assinaturas:", tn))
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(tab_ass)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def gerar_excel(do_mes, cols_ok, mes_nome, ano, tot):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        do_mes[cols_ok].to_excel(
            writer, sheet_name=mes_nome, index=False, startrow=7)
        ws = writer.sheets[mes_nome]
        ws["A1"] = "Paróquia SS. Trindade"
        ws["A2"] = "Livro de Ponto — Yolanda Facitela Clávio"
        ws["A3"] = "Pároco: Pe. Pasquale Peluso"
        ws["A4"] = "Secretária: Yolanda Facitela Clávio"
        ws["A5"] = f"Mês: {mes_nome} {ano}"
        ws["A6"] = f"Total de Horas: {tot//60}h {tot%60:02d}m"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"].alignment = Alignment(horizontal="center")
    return output.getvalue()
